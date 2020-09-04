import os
import pdfplumber
import re
from openpyxl import Workbook
from openpyxl import load_workbook
import time


class Drawing_Renamer_Tools:

    #defining constructor
    def __init__(self, filePath, main_file_list):
        self.file_path = filePath
        self.main_file_list = main_file_list


    def return_list_of_files(self):
        for file_name in os.listdir(self.file_path):
            self.main_file_list[file_name] = file_name
        return self.main_file_list


    def enter_beginningtext(self, beginning_text):
        for item in self.main_file_list:
            current_text = self.main_file_list[item]
            update_text = f'{beginning_text}{current_text}'
            self.main_file_list[item] = update_text
        return self.main_file_list


    def enter_endtext(self, end_text):
        for item in self.main_file_list:
            file_extension = item.split(".")[-1]
            current_text = self.main_file_list[item].split(".")[0]
            update_text = f'{current_text}{end_text}.{file_extension}'
            self.main_file_list[item] = update_text
        return self.main_file_list


    def enter_replace_text(self, replace_text_before, replace_text_after):
        for item in self.main_file_list:
            current_text = self.main_file_list[item]
            update_text = current_text.replace(replace_text_before, replace_text_after)
            self.main_file_list[item] = update_text
        return self.main_file_list


    def enter_titleblock_search(self, full_file_path, sample_drawing_number):
        # rename the file based on what is in the titleblock
        # Open the file and extract text with PyPDF2
        with pdfplumber.open(full_file_path) as pdf:
            full_text = ""
            pages = pdf.pages
            for i, pg in enumerate(pages):
                text = pages[i].extract_text()
                full_text += text
            # Convert the sample filename into regex search and search the extracted text
            output_text = ""
            regex_variable = "."
            for char in sample_drawing_number:
                if char == "-":
                    output_text += "-"
                elif char == "_":
                    output_text += "_"
                elif char == "(":
                    output_text += "\("
                elif char == ")":
                    output_text += "\)"
                else:
                    output_text += regex_variable
            regex_results = re.findall(output_text, full_text)
            number_of_regex_groups = len(regex_results) - 1
            best_guess_drawing_number = regex_results[number_of_regex_groups]
        return best_guess_drawing_number


    def rename_file(self, old_file_name, new_file_name):
        os.rename(old_file_name, new_file_name)



    def excel_list_export(self):
        drawing_list = Workbook()
        worksheet_1 = drawing_list.active
        drawing_list_file_path = self.file_path + "/" + "Drawing_List.xlsx"
        drawing_list.save(drawing_list_file_path)
        row_number = 1
        for key, values in self.main_file_list.items():
            file_name_no_extension = ""
            for char in key:
                if char != ".":
                    file_name_no_extension += char
                else:
                    break
            file_extension = key.split(".")[-1]
            worksheet_1.cell(row=row_number, column=1).value = file_name_no_extension
            worksheet_1.cell(row=row_number, column=2).value = file_extension
            worksheet_1.cell(row=row_number, column=3).value = file_name_no_extension
            row_number += 1
        drawing_list.save(drawing_list_file_path)



    def excel_list_import(self):
        drawing_list = Workbook()
        worksheet_1 = drawing_list.active
        drawing_list_file_path = self.file_path + "/" + "Drawing_List.xlsx"
        drawing_list = load_workbook(drawing_list_file_path)
        worksheet_1 = drawing_list.active
        row_number = worksheet_1.max_row
        row_number += 1  # to adjust the count as it starts from 0
        for i in range(1, row_number):
            old_file_name = worksheet_1.cell(row=i, column=1).value
            file_extension = worksheet_1.cell(row=i, column=2).value
            new_file_name = worksheet_1.cell(row=i, column=3).value
            old_file_name_plus_extension = old_file_name + "." + file_extension
            new_file_name_plus_extension = new_file_name + "." + file_extension
            self.main_file_list[old_file_name_plus_extension] = new_file_name_plus_extension
        return self.main_file_list


